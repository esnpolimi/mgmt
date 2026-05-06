import logging
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO

from django.conf import settings
from django.db.models import Case, Count, DecimalField, Sum, Value, When
from django.utils import timezone

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from treasury.models import Account, Transaction

try:
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover
    ZoneInfo = None

logger = logging.getLogger(__name__)

DRIVE_SCOPE = "https://www.googleapis.com/auth/drive"
EXCEL_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ROOT_FOLDER_NAME = "Treasury-Reports"
ACCOUNTS_FOLDER_NAME = "Casse"
TRANSACTIONS_FOLDER_NAME = "Transazioni"


class ReportDateError(ValueError):
    pass


def get_report_timezone():
    return ZoneInfo("Europe/Rome") if ZoneInfo else timezone.get_current_timezone()


def resolve_report_date(report_date, tz):
    if isinstance(report_date, datetime):
        return report_date.date()
    if isinstance(report_date, date):
        return report_date
    if isinstance(report_date, str) and report_date:
        try:
            return datetime.strptime(report_date, "%Y-%m-%d").date()
        except ValueError as exc:
            raise ReportDateError("Invalid date format. Use YYYY-MM-DD.") from exc

    now_local = timezone.localtime(timezone.now(), tz) if tz else timezone.localtime(timezone.now())
    return now_local.date()


def get_day_bounds(report_date, tz):
    start_dt = datetime.combine(report_date, time.min)
    if timezone.is_naive(start_dt):
        start_dt = timezone.make_aware(start_dt, tz)
    end_dt = start_dt + timedelta(days=1)
    return start_dt, end_dt


def _get_drive_service():
    credentials = service_account.Credentials.from_service_account_file(
        settings.GOOGLE_SERVICE_ACCOUNT_FILE,
        scopes=[DRIVE_SCOPE],
    )
    return build("drive", "v3", credentials=credentials)


def _find_or_create_drive_folder(service, folder_name, parent_id):
    escaped_folder_name = folder_name.replace("'", "\\'")
    query = (
        f"name='{escaped_folder_name}' and '{parent_id}' in parents "
        "and mimeType='application/vnd.google-apps.folder' and trashed=false"
    )
    results = service.files().list(
        q=query,
        spaces="drive",
        fields="files(id, name)",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()

    folders = results.get("files", [])
    if folders:
        return folders[0]["id"]

    folder_metadata = {
        "name": folder_name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id],
    }
    folder = service.files().create(
        body=folder_metadata,
        fields="id",
        supportsAllDrives=True,
    ).execute()
    return folder["id"]


def _find_file_id(service, folder_id, filename):
    escaped_filename = filename.replace("'", "\\'")
    query = f"name='{escaped_filename}' and '{folder_id}' in parents and trashed=false"
    results = service.files().list(
        q=query,
        spaces="drive",
        fields="files(id, name)",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()
    files = results.get("files", [])
    return files[0]["id"] if files else None


def _upload_excel(service, folder_id, filename, content_stream):
    content_stream.seek(0)
    media = MediaIoBaseUpload(content_stream, mimetype=EXCEL_MIMETYPE)
    existing_id = _find_file_id(service, folder_id, filename)
    if existing_id:
        service.files().update(
            fileId=existing_id,
            media_body=media,
            supportsAllDrives=True,
        ).execute()
        return existing_id, "updated"

    created = service.files().create(
        body={"name": filename, "parents": [folder_id]},
        media_body=media,
        fields="id",
        supportsAllDrives=True,
    ).execute()
    return created["id"], "created"


def _autosize_columns(worksheet):
    for column in worksheet.columns:
        max_len = 0
        col_letter = column[0].column_letter
        for cell in column:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        worksheet.column_dimensions[col_letter].width = min(max_len + 2, 50)


def _write_headers(worksheet, headers):
    header_font = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def _workbook_to_stream(workbook):
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def ensure_report_folders(service, parent_folder_id):
    root_folder_id = _find_or_create_drive_folder(service, ROOT_FOLDER_NAME, parent_folder_id)
    accounts_folder_id = _find_or_create_drive_folder(service, ACCOUNTS_FOLDER_NAME, root_folder_id)
    transactions_folder_id = _find_or_create_drive_folder(service, TRANSACTIONS_FOLDER_NAME, root_folder_id)
    return root_folder_id, accounts_folder_id, transactions_folder_id


def build_accounts_workbook(start_dt, end_dt, report_date):
    headers = [
        "Data",
        "Account ID",
        "Nome Cassa",
        "Status",
        "Saldo Iniziale",
        "Saldo Finale",
        "Totale Entrate",
        "Totale Uscite",
        "N Transazioni",
        "Note",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Casse"
    _write_headers(ws, headers)

    amount_field = DecimalField(max_digits=12, decimal_places=2)
    daily_totals = Transaction.objects.filter(
        created_at__gte=start_dt,
        created_at__lt=end_dt,
    ).values("account_id").annotate(
        total=Sum("amount"),
        total_in=Sum(
            Case(
                When(amount__gt=0, then="amount"),
                default=Value(Decimal("0.00")),
                output_field=amount_field,
            )
        ),
        total_out=Sum(
            Case(
                When(amount__lt=0, then="amount"),
                default=Value(Decimal("0.00")),
                output_field=amount_field,
            )
        ),
        count=Count("id"),
    )
    totals_by_account = {row["account_id"]: row for row in daily_totals}

    after_totals = Transaction.objects.filter(
        created_at__gte=end_dt,
    ).values("account_id").annotate(total=Sum("amount"))
    after_by_account = {row["account_id"]: row for row in after_totals}

    row_idx = 2
    report_date_str = report_date.strftime("%d/%m/%Y")
    for account in Account.objects.all().order_by("name"):
        totals = totals_by_account.get(account.id, {})
        total = totals.get("total") or Decimal("0.00")
        total_in = totals.get("total_in") or Decimal("0.00")
        total_out = totals.get("total_out") or Decimal("0.00")
        count = totals.get("count") or 0

        total_after = after_by_account.get(account.id, {}).get("total") or Decimal("0.00")
        saldo_finale = Decimal(str(account.balance or 0)) - Decimal(str(total_after))
        saldo_iniziale = saldo_finale - Decimal(str(total))

        ws.cell(row=row_idx, column=1, value=report_date_str)
        ws.cell(row=row_idx, column=2, value=account.id)
        ws.cell(row=row_idx, column=3, value=account.name)
        ws.cell(row=row_idx, column=4, value=account.status)
        ws.cell(row=row_idx, column=5, value=saldo_iniziale)
        ws.cell(row=row_idx, column=6, value=saldo_finale)
        ws.cell(row=row_idx, column=7, value=Decimal(str(total_in)))
        ws.cell(row=row_idx, column=8, value=abs(Decimal(str(total_out))))
        ws.cell(row=row_idx, column=9, value=count)
        ws.cell(row=row_idx, column=10, value="")

        for col in (5, 6, 7, 8):
            ws.cell(row=row_idx, column=col).number_format = "#,##0.00"

        row_idx += 1

    _autosize_columns(ws)
    return wb


def _executor_display(tx_obj):
    if tx_obj.executor and getattr(tx_obj.executor, "profile", None):
        return f"{tx_obj.executor.profile.name} {tx_obj.executor.profile.surname}"
    if tx_obj.executor:
        return getattr(tx_obj.executor, "email", "") or ""
    return ""


def build_transactions_workbook(start_dt, end_dt, tz):
    headers = [
        "ID",
        "Data/Ora",
        "Tipo",
        "Importo",
        "Cassa",
        "Descrizione",
        "Executor",
        "Subscription ID",
        "ESNcard ID",
        "Event Ref ID",
        "Receipt Link",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Transazioni"
    _write_headers(ws, headers)

    txs = Transaction.objects.filter(
        created_at__gte=start_dt,
        created_at__lt=end_dt,
    ).select_related(
        "account",
        "executor__profile",
    ).order_by("created_at", "id")

    row_idx = 2
    for tx in txs:
        local_dt = timezone.localtime(tx.created_at, tz) if tz else timezone.localtime(tx.created_at)
        ws.cell(row=row_idx, column=1, value=tx.id)
        ws.cell(row=row_idx, column=2, value=local_dt.strftime("%d/%m/%Y %H:%M:%S"))
        ws.cell(row=row_idx, column=3, value=tx.type)
        ws.cell(row=row_idx, column=4, value=Decimal(str(tx.amount)))
        ws.cell(row=row_idx, column=5, value=tx.account.name if tx.account else "")
        ws.cell(row=row_idx, column=6, value=tx.description)
        ws.cell(row=row_idx, column=7, value=_executor_display(tx))
        ws.cell(row=row_idx, column=8, value=tx.subscription_id or "")
        ws.cell(row=row_idx, column=9, value=tx.esncard_id or "")
        ws.cell(row=row_idx, column=10, value=tx.event_reference_manual_id or "")
        ws.cell(row=row_idx, column=11, value=tx.receipt_link or "")

        ws.cell(row=row_idx, column=4).number_format = "#,##0.00"
        row_idx += 1

    _autosize_columns(ws)
    return wb


def generate_accounts_report(report_date=None, tz=None, dry_run=False):
    tz = tz or get_report_timezone()
    report_date = resolve_report_date(report_date, tz)
    start_dt, end_dt = get_day_bounds(report_date, tz)
    filename = report_date.strftime("%d-%m-%Y") + ".xlsx"

    workbook = build_accounts_workbook(start_dt, end_dt, report_date)
    if dry_run:
        return {
            "filename": filename,
            "action": "dry-run",
            "report_date": report_date,
        }

    service = _get_drive_service()
    _, accounts_folder_id, _ = ensure_report_folders(service, settings.GOOGLE_DRIVE_FOLDER_ID)
    stream = _workbook_to_stream(workbook)
    file_id, action = _upload_excel(service, accounts_folder_id, filename, stream)
    return {
        "filename": filename,
        "file_id": file_id,
        "action": action,
        "report_date": report_date,
    }


def generate_transactions_report(report_date=None, tz=None, dry_run=False):
    tz = tz or get_report_timezone()
    report_date = resolve_report_date(report_date, tz)
    start_dt, end_dt = get_day_bounds(report_date, tz)
    filename = report_date.strftime("%d-%m-%Y") + ".xlsx"

    workbook = build_transactions_workbook(start_dt, end_dt, tz)
    if dry_run:
        return {
            "filename": filename,
            "action": "dry-run",
            "report_date": report_date,
        }

    service = _get_drive_service()
    _, _, transactions_folder_id = ensure_report_folders(service, settings.GOOGLE_DRIVE_FOLDER_ID)
    stream = _workbook_to_stream(workbook)
    file_id, action = _upload_excel(service, transactions_folder_id, filename, stream)
    return {
        "filename": filename,
        "file_id": file_id,
        "action": action,
        "report_date": report_date,
    }


def generate_daily_reports(report_date=None, tz=None, dry_run=False):
    tz = tz or get_report_timezone()
    report_date = resolve_report_date(report_date, tz)
    start_dt, end_dt = get_day_bounds(report_date, tz)

    accounts_wb = build_accounts_workbook(start_dt, end_dt, report_date)
    transactions_wb = build_transactions_workbook(start_dt, end_dt, tz)

    filename = report_date.strftime("%d-%m-%Y") + ".xlsx"
    if dry_run:
        return {
            "report_date": report_date,
            "accounts": {"filename": filename, "action": "dry-run"},
            "transactions": {"filename": filename, "action": "dry-run"},
        }

    service = _get_drive_service()
    _, accounts_folder_id, transactions_folder_id = ensure_report_folders(
        service, settings.GOOGLE_DRIVE_FOLDER_ID
    )

    accounts_stream = _workbook_to_stream(accounts_wb)
    acc_id, acc_action = _upload_excel(service, accounts_folder_id, filename, accounts_stream)

    transactions_stream = _workbook_to_stream(transactions_wb)
    tx_id, tx_action = _upload_excel(service, transactions_folder_id, filename, transactions_stream)

    return {
        "report_date": report_date,
        "accounts": {"filename": filename, "file_id": acc_id, "action": acc_action},
        "transactions": {"filename": filename, "file_id": tx_id, "action": tx_action},
    }
