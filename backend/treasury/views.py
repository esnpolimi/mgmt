import logging
from datetime import timedelta, datetime
from decimal import Decimal

import sentry_sdk
from django.core.exceptions import PermissionDenied, ObjectDoesNotExist
from django.db import transaction, IntegrityError
from django.db.models import Q
from django.http import HttpResponse
from django.utils.dateparse import parse_datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from rest_framework import serializers
from rest_framework.decorators import api_view, permission_classes
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.core.mail import send_mail

from events.models import Subscription, Event
from profiles.models import Profile
from treasury.models import Account, ESNcard, Settings, ReimbursementRequest
from treasury.models import Transaction
from treasury.serializers import TransactionViewSerializer, AccountDetailedViewSerializer, AccountEditSerializer, \
    AccountCreateSerializer, ESNcardEmissionSerializer, TransactionCreateSerializer, \
    ESNcardSerializer, AccountListViewSerializer, ReimbursementRequestSerializer, ReimbursementRequestViewSerializer, \
    TransactionUpdateSerializer
from users.models import User
from googleapiclient.errors import HttpError
from django.conf import settings
from django.utils import timezone
try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None

logger = logging.getLogger(__name__)


def user_is_board(user):
    return user.groups.filter(name="Board").exists()


def get_action_permissions(action, user):
    """
    Returns True if the user is allowed to perform the given action.
    """
    if action in ['reimburse_deposits', 'reimburse_quota']:
        return True  # user_is_board(user)
    if action == 'account_creation':
        return user_is_board(user)
    if action == 'account_detail_patch':
        return user.has_perm('treasury.change_account')
    if action == 'transaction_add':
        return user.has_perm('treasury.add_transaction')
    if action == 'transaction_detail_patch':
        return user.has_perm('treasury.change_transaction') or getattr(user, 'can_manage_casse', False)
    if action == 'transaction_detail_delete':
        return user.has_perm('treasury.delete_transaction') or getattr(user, 'can_manage_casse', False)
    if action == 'esncard_detail_patch':
        return user.has_perm('treasury.change_esncard')
    if action == 'reimbursement_request_detail_patch':
        return user_is_board(user)
    if action == 'reimbursement_request_detail_delete':
        return user_is_board(user) or user.has_perm('treasury.delete_reimbursementrequest')
    # Default: allow
    return True


def apply_transaction_filters(qs, request):
    """
    Reuse filtering logic between list and export.
    """
    search = request.GET.get('search', '').strip()
    if search:
        qs = qs.filter(
            Q(account__name__icontains=search) |
            Q(description__icontains=search) |
            (Q(executor__isnull=False) & (
                    Q(executor__profile__name__icontains=search) |
                    Q(executor__profile__surname__icontains=search)
            ))
        )
    event_id = request.GET.get('event')
    if event_id:
        qs = qs.filter(
            Q(subscription__event__id=event_id) |
            Q(event_reference_manual__id=event_id)
        )
    account_ids = request.GET.getlist('account')
    if account_ids:
        qs = qs.filter(account__id__in=account_ids)
    types = request.GET.getlist('type')
    if types:
        qs = qs.filter(type__in=types)
    date_from = request.GET.get('dateFrom')
    if date_from:
        qs = qs.filter(created_at__gte=date_from)
    date_to = request.GET.get('dateTo')
    if date_to:
        qs = qs.filter(created_at__lte=parse_datetime(date_to) + timedelta(days=1))
    return qs


# Endpoint for ESNcard emission.
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def esncard_emission(request):
    try:
        profile = Profile.objects.filter(id=request.data['profile_id']).first()
        latest_card = profile.latest_esncard if profile else None
        logger.info("Latest card for" + str(profile) + ": " + str(latest_card))
        esncard_serializer = ESNcardEmissionSerializer(data=request.data)
        if not esncard_serializer.is_valid():
            return Response(esncard_serializer.errors, status=400)

        with transaction.atomic():
            esncard = esncard_serializer.save()
            settings = Settings.get()
            if latest_card:
                if latest_card.is_valid:
                    amount = float(settings.esncard_lost_fee)
                    description = "Emissione ESNcard smarrita"
                else:
                    amount = float(settings.esncard_release_fee)
                    description = "Rinnovo ESNcard"
            else:
                amount = float(settings.esncard_release_fee)
                description = "Emissione ESNcard"

            # Create the transaction
            t = Transaction(
                type=Transaction.TransactionType.ESNCARD,
                esncard=esncard,
                account=esncard_serializer.validated_data['account_id'],
                executor=request.user,
                amount=amount,
                description=f"{description}: {esncard.profile.name} {esncard.profile.surname}"
            )
            t.save()

            # Return the newly created ESNcard with additional info
            response_data = esncard_serializer.data
            return Response(response_data, status=200)

    except PermissionDenied as e:
        return Response({'error': str(e)}, status=403)
    except IntegrityError as e:
        # Handle duplicate ESNcard number
        if 'number' in str(e) and ('unique' in str(e).lower() or 'duplicate' in str(e).lower()):
            return Response({
                'esncard_number': ['Questo numero ESNcard è già in uso.']
            }, status=400)
        logger.error(f"IntegrityError non gestito: {str(e)}")
        return Response({'error': 'Errore di integrità dei dati.'}, status=400)
    except (ObjectDoesNotExist, ValueError) as e:
        return Response({'error': str(e)}, status=400)
    except Exception as e:
        logger.error(str(e))
        sentry_sdk.capture_exception(e)
        return Response({'error': 'Errore interno del server.'}, status=500)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def esncard_detail(request, pk):
    try:
        esncard = ESNcard.objects.get(pk=pk)
        if request.method == 'PATCH':
            if not get_action_permissions('esncard_detail_patch', request.user):
                return Response({'error': 'Non autorizzato.'}, status=401)
            update_data = {}
            if 'number' in request.data:
                update_data['number'] = request.data['number']
            esncard_serializer = ESNcardSerializer(esncard, data=update_data, partial=True)
            if esncard_serializer.is_valid():
                esncard_serializer.save()
                return Response(esncard_serializer.data, status=200)  # Return updated data
            else:
                return Response(esncard_serializer.errors, status=400)
        else:
            return Response({'error': "Metodo non consentito"}, status=405)
    except ESNcard.DoesNotExist:
        return Response({'error': 'La ESNcard non esiste'}, status=400)
    except Exception as e:
        logger.error(str(e))
        sentry_sdk.capture_exception(e)
        return Response({'error': 'Errore interno del server.'}, status=500)


# Endpoint to retrieve ensncard fees
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def esncard_fees(_):
    try:
        # Get the settings and add fee information
        settings = Settings.get()
        response = Response({
            'esncard_release_fee': str(settings.esncard_release_fee),
            'esncard_lost_fee': str(settings.esncard_lost_fee)
        }, status=200)
        return response
    except Exception as e:
        logger.error(str(e))
        sentry_sdk.capture_exception(e)
        return Response({'error': 'Errore interno del server.'}, status=500)


#   Endpoint for adding a transaction
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def transaction_add(request):
    try:
        if not get_action_permissions('transaction_add', request.user):
            return Response({'error': 'Non autorizzato.'}, status=401)

        transaction_serializer = TransactionCreateSerializer(data=request.data, context={'request': request})
        if not transaction_serializer.is_valid():
            return Response(transaction_serializer.errors, status=400)

        transaction_type = transaction_serializer.validated_data['type']
        if transaction_type in [Transaction.TransactionType.DEPOSIT, Transaction.TransactionType.WITHDRAWAL]:
            if not request.user.has_perm('treasury.add_transaction'):
                return Response({'error': 'Non autorizzato.'}, status=401)

        try:
            tx = transaction_serializer.save()
        except ValueError as ve:
            return Response({'error': str(ve)}, status=400)
        except PermissionDenied as pe:
            return Response({'error': str(pe)}, status=403)

        # --- Email notification for manual deposit / withdrawal ---
        if 'localhost' in settings.SCHEME_HOST:
            print("Skipping email notification in localhost environment.")
        else:
            if transaction_type in [Transaction.TransactionType.DEPOSIT, Transaction.TransactionType.WITHDRAWAL]:
                try:
                    executor_profile = getattr(tx.executor, 'profile', None)
                    executor_name = f"{executor_profile.name} {executor_profile.surname}" if executor_profile else "N/D"
                    executor_email = executor_profile.email if executor_profile else (getattr(tx.executor, 'email', 'N/D'))
                    receipt_info = tx.receipt_link if tx.receipt_link else "Nessuna ricevuta caricata"
                    subject = f"Nuova transazione manuale: {'Deposito' if transaction_type == Transaction.TransactionType.DEPOSIT else 'Prelievo'}"
                    body = (
                        f"Tipo: {transaction_type}\n"
                        f"Importo: {tx.amount} EUR\n"
                        f"Cassa: {tx.account.name}\n"
                        f"Esecutore: {executor_name} ({executor_email})\n"
                        f"Descrizione: {tx.description}\n"
                        f"Data: {tx.created_at.strftime('%d/%m/%Y %H:%M')}\n"
                        f"Ricevuta: {receipt_info}\n"
                    )
                    send_mail(
                        subject=subject,
                        message=body,
                        from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', None),
                        recipient_list=['tesoriere@esnpolimi.it'],
                        fail_silently=True
                    )
                except Exception as mail_exc:
                    logger.warning(f"Errore invio email tesoriere (transazione manuale): {mail_exc}")

        return Response(status=200)
    except Exception as e:
        logger.error(str(e))
        sentry_sdk.capture_exception(e)
        return Response({'error': 'Errore interno del server.'}, status=500)


#   Endpoint to retrieve list of transactions
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def transactions_list(request):
    try:
        transactions = Transaction.objects.all().order_by('-created_at')
        transactions = apply_transaction_filters(transactions, request)
        # Support limit param for dashboard
        limit = request.GET.get('limit')
        if limit:
            try:
                limit = int(limit)
                transactions = transactions[:limit]
                serializer = TransactionViewSerializer(transactions, many=True)
                return Response({'results': serializer.data,
                                 'count': transactions.count() if hasattr(transactions, 'count') else len(
                                     transactions)})
            except ValueError:
                return Response({'error': 'Parametro limit non valido.'}, status=400)
        paginator = PageNumberPagination()
        paginator.page_size_query_param = 'page_size'
        page = paginator.paginate_queryset(transactions, request=request)
        serializer = TransactionViewSerializer(page, many=True)
        # Returns paginated response, use .results in frontend
        return paginator.get_paginated_response(serializer.data)
    except Exception as e:
        logger.error(str(e))
        sentry_sdk.capture_exception(e)
        return Response({'error': 'Errore interno del server.'}, status=500)


# Endpoint to retrive transaction details based on id
@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def transaction_detail(request, pk):
    try:
        transaction_obj = Transaction.objects.get(pk=pk)

        if request.method == 'GET':
            serializer = TransactionViewSerializer(transaction_obj)
            return Response(serializer.data, status=200)

        elif request.method == 'PATCH':
            if not get_action_permissions('transaction_detail_patch', request.user):
                return Response({'error': 'Non autorizzato.'}, status=401)

            # Executor handling: optional & nullable
            executor_raw = request.data.get('executor', '__not_provided__')
            executor_pk = transaction_obj.executor_id  # default: keep existing
            if executor_raw != '__not_provided__':
                if executor_raw in [None, '', 'null']:
                    executor_pk = None  # clear executor
                else:
                    try:
                        executor_pk = User.objects.get(profile__id=int(executor_raw)).pk
                    except (ValueError, User.DoesNotExist):
                        try:
                            executor_pk = User.objects.get(profile__email=executor_raw).pk
                        except User.DoesNotExist:
                            return Response({'error': 'Executor non valido.'}, status=400)

            update_payload = request.data.copy()  # may include: account, amount, description, receiptFile, remove_receipt
            # Remove executor from serializer payload
            update_payload.pop('executor', None)

            serializer = TransactionUpdateSerializer(
                transaction_obj,
                data=update_payload,
                partial=True,
                context={'request': request}
            )
            if not serializer.is_valid():
                return Response(serializer.errors, status=400)

            try:
                serializer.save()  # handles amount/account/description/receipt
            except HttpError as he:
                status = getattr(he, 'resp', None).status if getattr(he, 'resp', None) else None
                if status in (500, 502, 503, 504):
                    return Response({'error': 'Errore temporaneo durante l\'upload della ricevuta. Riprova.'}, status=503)
                return Response({'error': 'Errore upload ricevuta.'}, status=400)
            except Exception as e:
                logger.error(f"Errore salvataggio transazione (PATCH): {e}")
                return Response({'error': 'Errore durante l\'aggiornamento.'}, status=500)

            # Update executor separately (not handled by serializer)
            if executor_pk != transaction_obj.executor_id:
                transaction_obj.executor_id = executor_pk
                transaction_obj.save(update_fields=['executor'])

            return Response(status=200)

        elif request.method == 'DELETE':
            if not get_action_permissions('transaction_detail_delete', request.user):
                return Response({'error': 'Non autorizzato.'}, status=401)
            # Allow deletion only for specific transaction types
            list_deleteable = [
                Transaction.TransactionType.RIMBORSO_CAUZIONE,
                Transaction.TransactionType.REIMBURSEMENT,
                Transaction.TransactionType.RIMBORSO_QUOTA,
                Transaction.TransactionType.DEPOSIT,
                Transaction.TransactionType.WITHDRAWAL
            ]
            if transaction_obj.type in list_deleteable:
                if not (request.user.has_perm('treasury.delete_transaction') or request.user.can_manage_casse):
                    return Response({'error': 'Non autorizzato.'}, status=401)
                transaction_obj.delete()
                return Response(status=204)
            else:
                return Response({'error': 'Solo i Rimborsi possono essere eliminati manualmente.'}, status=400)
        else:
            return Response({'error': "Metodo non consentito"}, status=405)
    except Transaction.DoesNotExist:
        return Response({'error': 'Transazione non trovata.'}, status=404)
    except Exception as e:
        logger.error(str(e))
        sentry_sdk.capture_exception(e)
        return Response({'error': 'Errore interno del server.'}, status=500)


# Endpoint to retrieve all accounts
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def accounts_list(request):
    try:
        accounts = Account.objects.all().order_by('id')
        visible_accounts = [account for account in accounts if account.is_visible_to_user(request.user)]
        serializer = AccountListViewSerializer(visible_accounts, many=True, context={'request': request})
        return Response(serializer.data, status=200)
    except Exception as e:
        logger.error(str(e))
        sentry_sdk.capture_exception(e)
        return Response({'error': 'Errore interno del server.'}, status=500)


# Endpoint to create new account
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def account_creation(request):
    try:
        if not get_action_permissions('account_creation', request.user):
            return Response({'error': 'Non autorizzato.'}, status=401)

        account_serializer = AccountCreateSerializer(data=request.data)
        if not account_serializer.is_valid():
            return Response(account_serializer.errors, status=400)

        account_serializer.save(changed_by=request.user)
        return Response(status=200)
    except Exception as e:
        logger.error(str(e))
        sentry_sdk.capture_exception(e)
        return Response({'error': 'Errore interno del server.'}, status=500)


# Endpoint to retrieve account info / edit account
@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def account_detail(request, pk):
    try:
        account = Account.objects.get(pk=pk)

        if request.method == 'GET':
            serializer = AccountDetailedViewSerializer(account, context={'request': request})
            return Response(serializer.data, status=200)

        if request.method == 'PATCH':
            # Status-only toggle?
            incoming_keys = set(request.data.keys())
            status_only = incoming_keys.issubset({'status'}) and 'status' in incoming_keys

            # Permission: full edit requires existing permission; status-only allowed for casse managers
            is_casse_manager = (
                    request.user.can_manage_casse or
                    request.user.groups.filter(name__in=['Attivi', 'Board']).exists()
            )

            if not status_only:
                if not get_action_permissions('account_detail_patch', request.user):
                    return Response({'error': 'Non autorizzato.'}, status=401)
            else:
                if not is_casse_manager:
                    return Response({'error': 'Non autorizzato.'}, status=401)

            data = request.data.copy()
            data['changed_by'] = request.user
            serializer = AccountEditSerializer(account, data=data, partial=True)
            if not serializer.is_valid():
                return Response(serializer.errors, status=400)
            serializer.save()
            return Response(serializer.data, status=200)
        else:
            return Response({'error': "Metodo non consentito"}, status=405)
    except Account.DoesNotExist:
        return Response({'error': 'Account non trovato.'}, status=404)
    except Exception as e:
        logger.error(str(e))
        sentry_sdk.capture_exception(e)
        return Response({'error': 'Errore interno del server.'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reimbursement_request_creation(request):
    try:
        serializer = ReimbursementRequestSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            instance = serializer.save()
            # --- Email notification for reimbursement request creation ---
            if 'localhost' in settings.SCHEME_HOST:
                print("Skipping email notification in localhost environment.")
            else:
                try:
                    profile = getattr(instance.user, 'profile', None)
                    user_name = f"{profile.name} {profile.surname}" if profile else getattr(instance.user, 'email', 'N/D')
                    receipt_info = instance.receipt_link if instance.receipt_link else "Nessuna ricevuta caricata"
                    subject = f"Nuova richiesta di rimborso #{instance.id}"
                    body = (
                        f"Richiedente: {user_name}\n"
                        f"Importo: {instance.amount} EUR\n"
                        f"Metodo pagamento: {instance.payment}\n"
                        f"Descrizione: {instance.description}\n"
                        f"Data richiesta: {instance.created_at.strftime('%d/%m/%Y %H:%M')}\n"
                        f"Ricevuta: {receipt_info}\n"
                    )
                    send_mail(
                        subject=subject,
                        message=body,
                        from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', None),
                        recipient_list=['tesoriere@esnpolimi.it'],
                        fail_silently=True
                    )
                except Exception as mail_exc:
                    logger.warning(f"Errore invio email tesoriere (richiesta rimborso): {mail_exc}")

            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)
    except Exception as e:
        logger.error(str(e))
        sentry_sdk.capture_exception(e)
        return Response({'error': 'Errore interno del server.'}, status=500)


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def reimbursement_request_detail(request, pk):
    try:
        try:
            instance = ReimbursementRequest.objects.get(pk=pk)
        except ReimbursementRequest.DoesNotExist:
            return Response({'error': 'Richiesta di rimborso non trovata.'}, status=404)

        if request.method == 'GET':
            serializer = ReimbursementRequestViewSerializer(instance)
            return Response(serializer.data, status=200)

        elif request.method == 'PATCH':
            if not get_action_permissions('reimbursement_request_detail_patch', request.user):
                return Response({'error': 'Non autorizzato.'}, status=401)
            if request.user.has_perm('treasury.change_reimbursementrequest'):
                allowed_fields = {'description', 'receipt_link', 'account', 'amount'}
                data = {k: v for k, v in request.data.items() if k in allowed_fields}
                serializer = ReimbursementRequestSerializer(instance, data=data, partial=True,
                                                            context={'request': request})
                if serializer.is_valid():
                    try:
                        serializer.save()
                        # Reload instance to reflect changes (e.g., account deduction)
                        instance.refresh_from_db()
                        response_serializer = ReimbursementRequestViewSerializer(instance)
                        return Response(response_serializer.data, status=200)
                    except serializers.ValidationError as ve:
                        return Response(ve.detail, status=400)
                return Response(serializer.errors, status=400)
            else:
                return Response({'error': 'Non autorizzato.'}, status=401)

        elif request.method == 'DELETE':
            if not get_action_permissions('reimbursement_request_detail_delete', request.user):
                return Response({'error': 'Non autorizzato.'}, status=401)
            try:
                instance.delete()
                return Response(status=204)
            except Exception as del_exc:
                logger.error(f"Errore eliminazione richiesta rimborso #{pk}: {del_exc}")
                return Response({'error': 'Errore durante l\'eliminazione.'}, status=500)
        else:
            return Response({'error': "Metodo non consentito"}, status=405)
    except Exception as e:
        logger.error(str(e))
        sentry_sdk.capture_exception(e)
        return Response({'error': 'Errore interno del server.'}, status=500)


# Endpoint to retrieve list of reimbursement requests
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reimbursement_requests_list(request):
    try:
        requests = ReimbursementRequest.objects.all().order_by('-created_at')
        profile_id = request.GET.get('profile')
        if profile_id:
            try:
                profile_id_int = int(profile_id)
            except ValueError:
                return Response({'error': 'Parametro profile non valido.'}, status=400)
            # Allow only board or the profile owner
            req_profile_id = getattr(getattr(request.user, 'profile', None), 'id', None)
            if not user_is_board(request.user) and req_profile_id != profile_id_int:
                return Response({'error': 'Non autorizzato.'}, status=403)
            requests = requests.filter(user__profile__id=profile_id_int)
        search = request.GET.get('search', '').strip()
        if search:
            requests = requests.filter(
                Q(description__icontains=search) |
                Q(user__profile__name__icontains=search) |
                Q(user__profile__surname__icontains=search) |
                Q(user__profile__email__icontains=search)
            )
        # Filtering by payment (multi)
        payments = request.GET.getlist('payment')
        if payments:
            requests = requests.filter(payment__in=payments)
        # Filtering by dateFrom/dateTo
        date_from = request.GET.get('dateFrom')
        if date_from:
            requests = requests.filter(created_at__gte=date_from)
        date_to = request.GET.get('dateTo')
        if date_to:
            requests = requests.filter(created_at__lte=parse_datetime(date_to) + timedelta(days=1))
        # Support limit param for dashboard
        limit = request.GET.get('limit')
        if limit:
            try:
                limit = int(limit)
                requests = requests[:limit]
                serializer = ReimbursementRequestViewSerializer(requests, many=True)
                return Response({'results': serializer.data,
                                 'count': requests.count() if hasattr(requests, 'count') else len(requests)})
            except ValueError:
                return Response({'error': 'Parametro limit non valido.'}, status=400)
        paginator = PageNumberPagination()
        paginator.page_size_query_param = 'page_size'
        page = paginator.paginate_queryset(requests, request=request)
        serializer = ReimbursementRequestViewSerializer(page, many=True)
        return paginator.get_paginated_response(serializer.data)
    except Exception as e:
        logger.error(str(e))
        sentry_sdk.capture_exception(e)
        return Response({'error': 'Errore interno del server.'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reimburse_deposits(request):
    """
    Reimburse deposits for a list of subscriptions.
    Expects: { "event": <event_id>, "subscription_ids": [<id1>, <id2>, ...], "account": <account_id>, "notes": <optional> }
    """
    try:
        if not get_action_permissions('reimburse_deposits', request.user):
            return Response({'error': 'Non autorizzato.'}, status=401)

        event_id = request.data.get('event')
        subscription_ids = request.data.get('subscription_ids', [])
        account_id = request.data.get('account')
        notes = request.data.get('notes', '')

        if not event_id or not subscription_ids or not account_id:
            return Response({'error': 'Dati mancanti.'}, status=400)

        event = Event.objects.get(id=event_id)
        account = Account.objects.get(id=account_id)
        subscriptions = Subscription.objects.filter(id__in=subscription_ids, event=event)

        # --- Restrict reimbursement if event.reimbursements_by_organizers_only is True ---
        if getattr(event, 'reimbursements_by_organizers_only', False):
            is_board = user_is_board(request.user)
            organizer_ids = set(event.organizers.values_list('profile_id', flat=True))
            user_profile_id = getattr(getattr(request.user, 'profile', None), 'id', None)
            if not (is_board or (user_profile_id and user_profile_id in organizer_ids)):
                return Response({'error': 'Solo gli organizzatori o Board possono rimborsare per questo evento.'}, status=403)

        if not subscriptions.exists():
            return Response({'error': 'Nessuna iscrizione valida trovata.'}, status=400)

        deposit_amount = event.deposit or Decimal('0.00')

        with transaction.atomic():
            account_locked = Account.objects.select_for_update().get(pk=account.pk)
            created = []
            for sub in subscriptions:
                if Transaction.objects.filter(subscription=sub,
                                              type=Transaction.TransactionType.RIMBORSO_CAUZIONE).exists():
                    continue
                cauzione_tx = Transaction.objects.filter(subscription=sub,
                                                         type=Transaction.TransactionType.CAUZIONE).first()
                if not cauzione_tx:
                    # Fix: handle external subscriptions gracefully
                    sub_name = None
                    if sub.profile:
                        sub_name = f"{sub.profile.name} {sub.profile.surname}"
                    elif sub.external_name:
                        sub_name = sub.external_name
                    else:
                        sub_name = "Esterno"
                    return Response({'error': f'Nessuna cauzione rimborsabile trovata per {sub_name}'}, status=400)
                if account_locked.status == "closed":
                    return Response({'error': 'La cassa è chiusa.'}, status=400)
                if account_locked.balance < deposit_amount:
                    return Response({'error': 'Saldo cassa insufficiente.'}, status=400)
                # Fix: handle external subscriptions for description
                if sub.profile:
                    sub_name = f"{sub.profile.name} {sub.profile.surname}"
                elif sub.external_name:
                    sub_name = sub.external_name
                else:
                    sub_name = "Esterno"
                tx = Transaction.objects.create(
                    type=Transaction.TransactionType.RIMBORSO_CAUZIONE,
                    subscription=sub,
                    executor=request.user,
                    account=account_locked,
                    amount=-deposit_amount,
                    description=f"Rimborso cauzione {sub_name} - {event.name}" + (f" - {notes}" if notes else "")
                )
                created.append(tx)
            # Return the created transactions
            serializer = TransactionViewSerializer(created, many=True)
            return Response(serializer.data, status=201)
    except Exception as e:
        logger.error(str(e))
        sentry_sdk.capture_exception(e)
        return Response({'error': 'Errore interno del server.'}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reimbursable_deposits(request):
    """
    Returns a list of subscriptions for a given event and list that are eligible for deposit reimbursement.
    Query params: event=<event_id>&list=<list_id>
    """
    try:
        event_id = request.GET.get('event')
        list_id = request.GET.get('list')
        if not event_id or not list_id:
            return Response({'error': 'Evento e Lista richiesti.'}, status=400)

        event = Event.objects.get(id=event_id)

        # Subscriptions in this list, with a paid cauzione transaction, not yet reimbursed
        subs = Subscription.objects.filter(event=event, list__id=list_id)
        result = []
        for sub in subs:
            deposit_tx = Transaction.objects.filter(subscription=sub, type=Transaction.TransactionType.CAUZIONE).first()
            reimbursed = Transaction.objects.filter(subscription=sub,
                                                    type=Transaction.TransactionType.RIMBORSO_CAUZIONE).exists()
            if deposit_tx and not reimbursed:
                result.append({
                    "id": sub.pk,
                    "profile_id": sub.profile.id,
                    "profile_name": f"{sub.profile.name} {sub.profile.surname}",
                    "account_name": deposit_tx.account.name if deposit_tx.account else None
                })
        return Response(result, status=200)
    except Exception as e:
        logger.error(str(e))
        sentry_sdk.capture_exception(e)
        return Response({'error': 'Errore interno del server.'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reimburse_quota(request):
    """
    Reimburse the main event cost ("Quota") for a subscription.
    Expects: { "event": <event_id>, "subscription_id": <id>, "account": <account_id>, "notes": <optional>, "include_services": <optional bool> }
    """
    try:
        if not get_action_permissions('reimburse_quota', request.user):
            return Response({'error': 'Non autorizzato.'}, status=401)

        event_id = request.data.get('event')
        subscription_id = request.data.get('subscription_id')
        account_id = request.data.get('account')
        notes = request.data.get('notes', '')
        include_services_raw = request.data.get('include_services', False)

        def _as_bool(val):
            if isinstance(val, bool):
                return val
            if val is None:
                return False
            return str(val).strip().lower() in ['1', 'true', 'yes', 'y', 'on']

        include_services = _as_bool(include_services_raw)

        if not event_id or not subscription_id or not account_id:
            return Response({'error': 'Dati mancanti.'}, status=400)

        event = Event.objects.get(id=event_id)
        account = Account.objects.get(id=account_id)

        sub = Subscription.objects.select_related('event', 'list').get(id=subscription_id)

        # Ensure the subscription's list is actually attached to the requested event (shared lists scenario)
        if event not in sub.list.events.all():
            return Response({'error': 'La sottoscrizione non appartiene alle liste di questo evento.'}, status=400)

        # Use the subscription's owning event for quota metadata (cost, name, etc.)
        sub_event = sub.event

        # --- Restrict reimbursement if event.reimbursements_by_organizers_only is True ---
        if getattr(event, 'reimbursements_by_organizers_only', False):
            is_board = user_is_board(request.user)
            organizer_ids = set(event.organizers.values_list('profile_id', flat=True))
            user_profile_id = getattr(getattr(request.user, 'profile', None), 'id', None)
            if not (is_board or (user_profile_id and user_profile_id in organizer_ids)):
                return Response({'error': 'Solo gli organizzatori o Board possono rimborsare per questo evento.'}, status=403)

        # Only allow if event is not free and subscription has a paid transaction
        if not sub_event.cost or float(sub_event.cost) <= 0:
            return Response({'error': 'L\'evento è gratuito, nessuna quota da rimborsare.'}, status=400)
        if not Transaction.objects.filter(subscription=sub, type=Transaction.TransactionType.SUBSCRIPTION).exists():
            return Response({'error': 'La quota può essere rimborsata solo se è stato effettuato il pagamento.'},
                            status=400)

        quota_already_reimbursed = Transaction.objects.filter(
            subscription=sub,
            type=Transaction.TransactionType.RIMBORSO_QUOTA
        ).exists()
        if quota_already_reimbursed and not include_services:
            return Response({'error': 'Quota già rimborsata.'}, status=400)

        quota_tx = Transaction.objects.filter(subscription=sub, type=Transaction.TransactionType.SUBSCRIPTION).first()
        if not quota_tx:
            return Response({'error': 'Nessun pagamento quota trovato per questa iscrizione.'}, status=400)

        if account.status == "closed":
            return Response({'error': 'La cassa è chiusa.'}, status=400)

        services_total = Decimal('0')
        if include_services:
            selected_services = sub.selected_services or []
            if not selected_services:
                return Response({'error': 'Nessun servizio selezionato per questa iscrizione.'}, status=400)
            if not Transaction.objects.filter(subscription=sub, type=Transaction.TransactionType.SERVICE).exists():
                return Response({'error': 'I servizi possono essere rimborsati solo se è stato effettuato il pagamento.'},
                                status=400)
            if Transaction.objects.filter(subscription=sub, type=Transaction.TransactionType.RIMBORSO_SERVICE).exists():
                return Response({'error': 'Servizi già rimborsati.'}, status=400)
            for s in selected_services:
                try:
                    price = Decimal(str(s.get('price_at_purchase') or s.get('price') or 0))
                except Exception:
                    price = Decimal('0')
                try:
                    qty = int(s.get('quantity') or 1)
                except Exception:
                    qty = 1
                if qty > 0:
                    services_total += (price * qty)
            if services_total <= 0:
                return Response({'error': 'Importo servizi non valido.'}, status=400)

        quota_amount = Decimal(str(sub_event.cost)) if not quota_already_reimbursed else Decimal('0')
        total_refund = quota_amount + services_total
        if account.balance < total_refund:
            return Response({'error': 'Saldo cassa insufficiente.'}, status=400)

        # Fix: handle external subscriptions for description
        if sub.profile:
            sub_name = f"{sub.profile.name} {sub.profile.surname}"
        elif sub.external_name:
            sub_name = sub.external_name
        else:
            sub_name = "Esterno"

        with transaction.atomic():
            quota_tx = None
            if not quota_already_reimbursed:
                quota_tx = Transaction.objects.create(
                    type=Transaction.TransactionType.RIMBORSO_QUOTA,
                    subscription=sub,
                    executor=request.user,
                    account=account,
                    amount=-sub_event.cost,
                    description=f"Rimborso quota {sub_name} - {sub_event.name}" + (f" - {notes}" if notes else "")
                )
            services_tx = None
            if include_services and services_total > 0:
                services_tx = Transaction.objects.create(
                    type=Transaction.TransactionType.RIMBORSO_SERVICE,
                    subscription=sub,
                    executor=request.user,
                    account=account,
                    amount=-services_total,
                    description=f"Rimborso servizi {sub_name} - {sub_event.name}" + (f" - {notes}" if notes else "")
                )
            payload = {
                'quota': TransactionViewSerializer(quota_tx).data if quota_tx else None,
                'services': TransactionViewSerializer(services_tx).data if services_tx else None
            }
            return Response(payload, status=201)
    except Exception as e:
        logger.error(str(e))
        sentry_sdk.capture_exception(e)
        return Response({'error': 'Errore interno del server.'}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def transactions_export(request):
    """
    Export filtered transactions with columns:
    Esecuzione | Esecuzione | Attività | Descrizione | Importo | Cassa | Commenti | Descrizione (gestionale)
    """

    # Helper to compute (possibly adjusted) amount, with narrowed exception handling.
    def compute_amount(tx_obj):
        amt = float(tx_obj.amount) if tx_obj.amount is not None else None
        if tx_obj.type == Transaction.TransactionType.ESNCARD and (amt is None or amt == 0):
            try:
                settings_obj = Settings.get()
                desc_lower = (tx_obj.description or "").lower()
                if "smarrita" in desc_lower:
                    amt = float(settings_obj.esncard_lost_fee)
                else:
                    amt = float(settings_obj.esncard_release_fee)
            except (Settings.DoesNotExist, AttributeError, ValueError) as calc_exc:
                logger.error(f"Errore calcolo importo ESNcard (tx id={tx_obj.id}): {calc_exc}")
                return None
        return amt

    txs_qs = Transaction.objects.all().select_related(
        'account',
        'subscription__event',
        'subscription__profile',
        'esncard__profile',
        'event_reference_manual',
        'executor__profile'
    ).order_by('-created_at')
    txs_qs = apply_transaction_filters(txs_qs, request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Bilancio"

    # Updated headers: inserted computed "Descrizione" and renamed original.
    headers = ["Esecuzione", "Esecuzione", "Attività", "Descrizione", "Importo", "Cassa", "Commenti", "Descrizione (gestionale)"]
    header_font = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    def build_attivita(tx_obj):
        if tx_obj.type == Transaction.TransactionType.ESNCARD:
            return "Quota Associativa"
        if tx_obj.type == Transaction.TransactionType.DEPOSIT:
            return "Deposito"
        if tx_obj.type == Transaction.TransactionType.WITHDRAWAL:
            return "Prelievo"
        if tx_obj.type == Transaction.TransactionType.REIMBURSEMENT:
            return "Richiesta Rimborso"
        ev = None
        if getattr(tx_obj, 'subscription_id', None) and getattr(tx_obj.subscription, 'event', None):
            ev = tx_obj.subscription.event
        elif getattr(tx_obj, 'event_reference_manual_id', None) and tx_obj.event_reference_manual:
            ev = tx_obj.event_reference_manual
        if not ev:
            return ""
        return ev.name or ""

    def _academic_year(date_obj):
        # Academic year starts Aug 1.
        y = date_obj.year
        if date_obj.month >= 9:
            start = y % 100
        else:
            start = (y - 1) % 100
        end = (start + 1) % 100
        return f"{start:02d}/{end:02d}"

    def build_descrizione(tx_obj):
        # New computed description column.
        if tx_obj.type == Transaction.TransactionType.ESNCARD:
            return f"Quota Associativa {_academic_year(tx_obj.created_at)}"
        ev = None
        if getattr(tx_obj, 'subscription_id', None) and getattr(tx_obj.subscription, 'event', None):
            ev = tx_obj.subscription.event
        elif getattr(tx_obj, 'event_reference_manual_id', None) and tx_obj.event_reference_manual:
            ev = tx_obj.event_reference_manual
        if not ev:
            return ""
        ev_date = getattr(ev, 'date', None) or getattr(ev, 'start_date', None) or getattr(ev, 'start', None)
        if not ev_date:
            return ""
        try:
            base_date = ev_date.strftime('%d/%m')
        except Exception:
            return ""
        mapping = {
            Transaction.TransactionType.SUBSCRIPTION: "iscrizione",
            Transaction.TransactionType.CAUZIONE: "cauzione",
            Transaction.TransactionType.RIMBORSO_QUOTA: "rimborso iscrizione",
            Transaction.TransactionType.RIMBORSO_CAUZIONE: "rimborso cauzione",
            Transaction.TransactionType.SERVICE: "servizio",
            Transaction.TransactionType.RIMBORSO_SERVICE: "rimborso servizio"
        }
        label = mapping.get(tx_obj.type)
        if not label:
            return ""
        return f"{base_date} - {label}"

    def build_commenti(tx_obj):
        if getattr(tx_obj, 'subscription_id', None) and getattr(tx_obj.subscription, 'profile', None):
            p = tx_obj.subscription.profile
            return f"{p.name} {p.surname}"
        if getattr(tx_obj, 'esncard_id', None) and getattr(tx_obj.esncard, 'profile', None):
            p = tx_obj.esncard.profile
            return f"{p.name} {p.surname}"
        if getattr(tx_obj, 'executor_id', None) and getattr(tx_obj.executor, 'profile', None):
            p = tx_obj.executor.profile
            return f"{p.name} {p.surname}"
        return ""

    row_idx = 2
    for tx in txs_qs:
        amount_val = compute_amount(tx)
        # Localize to Italian timezone (handles DST)
        if ZoneInfo is not None:
            local_dt = timezone.localtime(tx.created_at, ZoneInfo('Europe/Rome'))
        else:
            # Fallback to current timezone if zoneinfo unavailable
            local_dt = timezone.localtime(tx.created_at)
        ws.cell(row=row_idx, column=1, value=local_dt.strftime('%d/%m/%Y %H:%M:%S'))
        ws.cell(row=row_idx, column=2, value=local_dt.strftime('%d/%m/%Y %H:%M:%S'))
        ws.cell(row=row_idx, column=3, value=build_attivita(tx))
        ws.cell(row=row_idx, column=4, value=build_descrizione(tx))
        amt_cell = ws.cell(row=row_idx, column=5, value=amount_val)
        amt_cell.number_format = '#,##0.00 €'
        ws.cell(row=row_idx, column=6, value=tx.account.name if tx.account else '')
        ws.cell(row=row_idx, column=7, value=build_commenti(tx))
        ws.cell(row=row_idx, column=8, value=tx.description)
        row_idx += 1

    # Auto width
    for column in ws.columns:
        max_len = 0
        col_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    from io import BytesIO
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    event_id = request.GET.get('event')
    if event_id:
        try:
            ev_for_name = Event.objects.get(id=event_id)
            base = f"Bilancio_Evento_{ev_for_name.name.replace(' ', '_')}"
        except Event.DoesNotExist:
            base = "Bilancio_Evento"
    else:
        base = "Bilancio_Transazioni"

    filename = f"{base}_{datetime.now().strftime('%d%m%Y_%H%M%S')}.xlsx"
    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}'
    return response
