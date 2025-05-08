"""
    Use me this way, after entering the main folder of the project/backend:
    python manage.py importProfilesFromCSV "C:\\path\to\file.csv" --dry-run
    (the csv is considered exported from the old gest's db)
    (use the option --dry-run to NOT create the objects!)
    (use the option --export to generate an Excel preview file)
"""
import re
import os

from django.contrib.auth.models import Group
from django.core.management.base import BaseCommand
import csv
from datetime import datetime
from django.utils import timezone
from django.db import transaction
from profiles.models import Profile, Document
from treasury.models import ESNcard
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from users.models import User
from utils.country_prefix_utils import load_country_data, map_country_to_code, get_prefix_by_country_code


class Command(BaseCommand):
    help = 'Import profiles from a CSV file'
    country_codes = []
    AA = None
    status = None

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the CSV file')
        parser.add_argument('--dry-run', action='store_true', help='Run without saving to database')
        parser.add_argument('--export', nargs='?', type=str, const='./profiles_export.xlsx',
                            help='Enable export. Optionally specify path (defaults to profiles_export.xlsx)')

    def handle(self, *args, **options):
        dry_run = options.get('dry_run', False)
        file_path = options['file_path']
        export_path = options.get('export')

        # Add debug output for export path
        if export_path:
            self.stdout.write(f"Will export to: {os.path.abspath(export_path)}")

        # Load country codes and store in instance variable
        self.country_codes = load_country_data()
        if not self.country_codes:
            self.stdout.write(self.style.ERROR("Failed to load country codes"))
            return

        result = self.import_profiles_from_csv(file_path, export_path, dry_run)

        self.stdout.write(self.style.SUCCESS(f"Successfully imported {result['success_count']} profiles"))
        if result['export_path']:
            self.stdout.write(self.style.SUCCESS(f"Exported data to {result['export_path']}"))
        else:
            if export_path:
                self.stdout.write(self.style.ERROR(f"Failed to export data"))

        if result['error_count'] > 0:
            self.stdout.write(self.style.WARNING(f"Failed to import {result['error_count']} profiles"))
            for error in result['errors']:
                self.stdout.write(self.style.ERROR(error))

    def import_profiles_from_csv(self, file_path, export_path, dry_run=False):
        """
        Import profiles from a CSV file.

        Args:
            file_path: Path to the CSV file
            export_path: Path to export Excel file
            dry_run: Whether to commit changes to the database
        """
        # Initialize counters and collectors
        success_count = 0
        error_count = 0
        errors = []
        export_data = []

        if not self.country_codes:
            errors.append("Failed to load country codes")
            return {'success_count': 0, 'error_count': 1, 'errors': errors, 'export_path': None}

        # Document type mapping
        document_type_mapping = {
            'PA': Document.Type.PASSPORT,
            'ID': Document.Type.NATIONAL_ID,
        }

        with open(file_path, 'r', encoding='utf-8', errors='replace') as csv_file:
            # First detect actual headers
            csv_reader = csv.reader(csv_file, delimiter=',')
            headers = next(csv_reader)
            self.stdout.write(f"CSV headers: {headers}")

            # Reset file pointer and create DictReader
            csv_file.seek(0)
            csv_reader = csv.DictReader(csv_file, delimiter=',')

            # Process each row
            for row in csv_reader:
                try:
                    # Process profile data - do this outside the transaction
                    profile, document, esncard = self._process_row(row, document_type_mapping)
                    if not profile:
                        continue

                    # Collect data for export regardless of dry run
                    profile_data = self._create_export_data(profile, document, esncard)
                    export_data.append(profile_data)

                    # Only save to database if not dry run
                    if not dry_run:
                        with transaction.atomic():
                            profile.save()
                            if document:
                                document.profile = profile
                                document.save()
                            if esncard:
                                esncard.profile = profile
                                esncard.save()
                            if profile.is_esner:
                                user = User.objects.create_user(
                                    profile=profile,
                                    password=None
                                )
                                user.is_active = False  # Will be activated upon verification
                                if row['status'] == 'Aspirante':
                                    group, created = Group.objects.get_or_create(name="Aspiranti")
                                elif row['status'] == 'Associato':
                                    group, created = Group.objects.get_or_create(name="Attivi")
                                user.groups.add(group)

                    success_count += 1
                except Exception as e:
                    error_count += 1
                    errors.append(f"Error processing row for {row.get('email', 'unknown')}: {str(e)}")

        if dry_run:
            self.stdout.write(self.style.WARNING("Dry run, no changes saved to database"))

        # Create Excel export after processing all rows - outside any transaction
        export_file_path = None
        if export_data and export_path:
            try:
                export_file_path = self._create_excel_export(export_data, export_path)
            except Exception as e:
                errors.append(f"Error creating Excel export: {str(e)}")

        return {
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors,
            'export_path': export_file_path
        }

    def _process_row(self, row, document_type_mapping):
        """Process a single CSV row and return profile, document, and ESNcard objects."""
        if ((self.AA is not None and row['AA'] != self.AA) or
                (self.status is not None and row['status'] != self.status)):
            return None, None, None

        if 'nome' in row:
            self.stdout.write(f"Processing {row['nome']} {row['cognome']}")

        # Parse birthdate
        birthdate = None
        try:
            if 'data_nascita' in row and row['data_nascita']:
                birthdate = datetime.strptime(row['data_nascita'], '%d/%m/%Y').date()
        except ValueError:
            pass

        # Get country code
        country_name = row.get('nazione', '')
        country_code = map_country_to_code(country_name, self.country_codes)
        if not country_code:
            self.stdout.write(f"Failed to map country code for {country_name}")
            return None, None, None

        # Parse phone numbers
        phone_prefix, phone_number = self.parse_phone_number(row.get('telefono', ''), country_name)

        # Parse WhatsApp numbers, matricola expiration only for Erasmus
        if row['status'] == 'Erasmus':
            whatsapp_prefix, whatsapp_number = self.parse_phone_number(row.get('whatsapp', ''), country_name)
            # Use phone number as WhatsApp if WhatsApp not provided
            if not whatsapp_prefix and not whatsapp_number and (phone_prefix or phone_number):
                whatsapp_prefix = phone_prefix
                whatsapp_number = phone_number
        else:
            whatsapp_prefix, whatsapp_number = None, None

        # Parse matricola and person code
        matricola_number = self._parse_matricola(row.get('matricola', ''))
        person_code = self._parse_person_code(row.get('codice_persona', ''))

        # Parse registration date
        try:
            created_at = datetime.strptime(row.get('data_iscrizione', ''), '%Y-%m-%d %H:%M:%S')
        except ValueError:
            created_at = timezone.now()

        # Determine period end date for matricola expiration (only for Erasmus)
        if row['status'] == 'Erasmus':
            periodo = row.get('periodo_permanenza', '')
            end_date = '2025-01-31' if periodo == '1' else '2025-07-31'  # Assumed Erasmus of AA 2024
            matricola_expiration = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            matricola_expiration = None

        # Create Profile
        profile = Profile(
            email=row.get('email', ''),
            name=row.get('nome', ''),
            surname=row.get('cognome', ''),
            birthdate=birthdate,
            country=country_code,
            phone_prefix=phone_prefix,
            phone_number=phone_number,
            whatsapp_prefix=whatsapp_prefix,
            whatsapp_number=whatsapp_number,
            person_code=person_code,
            domicile=row.get('indirizzo_residenza', ''),
            matricola_number=matricola_number,
            matricola_expiration=matricola_expiration,
            email_is_verified=True,
            created_at=created_at,
            is_esner=row.get('status', '') != 'Erasmus'
        )

        # Create Document if document data is available
        document = self._create_document(row, document_type_mapping)

        # Create ESNcard if card number is available
        esncard = self._create_esncard(row, created_at)

        return profile, document, esncard

    def parse_phone_number(self, phone_str, country_name):
        """Parse phone number with country context."""
        if not phone_str or phone_str == 'NULL' or phone_str == '--':
            return None, None

        if phone_str.startswith('+'):
            # Try to match country code from our database
            for country in self.country_codes:
                dial_code = country['dial']
                if phone_str.startswith(dial_code):
                    number_part = phone_str[len(dial_code):]
                    return dial_code, int(number_part) if number_part.isdigit() else None

            # If we couldn't match a specific code, use a generic approach
            match = re.match(r'^\+(\d{1,4})(\d+)$', phone_str)
            if match:
                prefix = '+' + match.group(1)
                number = int(match.group(2))
                return prefix, number
            else:
                # If we can't match, just return with generic + prefix
                return '+', int(phone_str[1:]) if phone_str[1:].isdigit() else None

        # Handle 00XX format (international)
        elif phone_str.startswith('00'):
            # Convert 00 to + format
            plus_format = '+' + phone_str[2:]
            # Try again with + format
            return self.parse_phone_number(plus_format, country_name)

        # If we have a country name, try to use its dial code
        elif country_name:
            dial_code = get_prefix_by_country_code(country_name, self.country_codes)
            if dial_code:
                return dial_code, int(phone_str) if phone_str.isdigit() else None

        # Handle local numbers (no prefix)
        try:
            return None, int(phone_str) if phone_str.isdigit() else None
        except ValueError:
            return None, None

    @staticmethod
    def _parse_matricola(matricola):
        """Parse matricola number."""
        try:
            if matricola != 'NULL' and matricola.isdigit() and len(matricola) == 6:
                return int(matricola)
        except (ValueError, AttributeError):
            pass
        return None

    @staticmethod
    def _parse_person_code(person_code):
        """Parse person code."""
        try:
            if person_code != 'NULL' and person_code.isdigit() and len(person_code) == 8:
                return int(person_code)
        except (ValueError, AttributeError):
            pass
        return None

    @staticmethod
    def _create_document(row, document_type_mapping):
        """Create document object if data is available."""
        documento = row.get('documento', '')
        tipo_documento = row.get('tipo_documento', '')

        if not (documento and documento != 'NULL' and tipo_documento in document_type_mapping):
            return None

        try:
            doc_date = row.get('data_documento', '')
            doc_expiration = datetime.strptime(doc_date, '%d/%m/%Y').date() if doc_date else None
        except ValueError:
            doc_expiration = datetime.strptime('2030-01-01', '%Y-%m-%d').date()

        return Document(
            type=document_type_mapping.get(tipo_documento, Document.Type.OTHER),
            number=documento,
            expiration=doc_expiration
        )

    @staticmethod
    def _create_esncard(row, created_at):
        """Create ESNcard object if data is available."""
        card_number = row.get('ESN_card', '')
        if not (card_number and card_number != 'NULL'):
            return None

        return ESNcard(
            number=card_number,
            created_at=created_at
        )

    @staticmethod
    def _create_export_data(profile, document, esncard):
        """Create a dictionary of data to export."""
        return {
            'Name': profile.name,
            'Surname': profile.surname,
            'Email': profile.email,
            'Matricola': profile.matricola_number,
            'Matricola Expires': profile.matricola_expiration,
            'Person Code': profile.person_code,
            'Phone': f"{profile.phone_prefix or ''} - {profile.phone_number or ''}" if (profile.phone_prefix or profile.phone_number) else None,
            'WhatsApp': f"{profile.whatsapp_prefix or ''} - {profile.whatsapp_number or ''}" if (profile.whatsapp_prefix or profile.whatsapp_number) else None,
            'Birthdate': profile.birthdate,
            'Country': str(profile.country) if profile.country else None,
            'Domicile': profile.domicile,
            'Created At': profile.created_at,
            'Document Type': getattr(document, 'type', None),
            'Document Number': getattr(document, 'number', None),
            'Document Expiration': getattr(document, 'expiration', None),
            'ESNcard Number': getattr(esncard, 'number', None),
            'ESNcard Created': getattr(esncard, 'created_at', None)
        }

    def _create_excel_export(self, export_data, export_path):
        """Create Excel export file."""
        if not export_data:
            self.stdout.write(self.style.WARNING("No data to export"))
            return None

        try:
            # Make sure directory exists
            export_dir = os.path.dirname(os.path.abspath(export_path))
            if export_dir and not os.path.exists(export_dir):
                os.makedirs(export_dir, exist_ok=True)

            # Create a workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Profiles"

            # Add headers
            headers = list(export_data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            # Add data
            for row_idx, data_row in enumerate(export_data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = data_row.get(header)
                    # Convert None to empty string
                    if value is None:
                        value = ''
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto adjust column width
            for col_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                for row_idx in range(1, len(export_data) + 2):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = max_length + 2
                ws.column_dimensions[col_letter].width = adjusted_width

            # Save workbook with absolute path for clarity
            abs_path = os.path.abspath(export_path)
            self.stdout.write(f"Attempting to save Excel file to {abs_path}")
            wb.save(abs_path)

            # Verify file was created
            if os.path.exists(abs_path):
                self.stdout.write(f"Excel file created successfully at {abs_path}")
                return abs_path
            else:
                self.stdout.write(self.style.ERROR(f"Excel file wasn't created at {abs_path}"))
                return None

        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error saving Excel file: {str(e)}"))
            return None
